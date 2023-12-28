from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec
import pytest
import openpyxl
from constants import globalConstants as c


#Bir önceki ödevdeki işlemler, verileri Excel dosyasından alarak yeniden yapılmıştır.

class Test_LoginClass:
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get("https://www.saucedemo.com")
        self.driver.maximize_window() 

    def teardown_method(self): 
        self.driver.quit()

    #case 1-Kullanıcı adı ve şifre alanları boş geçildiğinde uyarı mesajı olarak "Epic sadface: Username is required" gösterilmelidir.
    @pytest.mark.parametrize("username, password", [("","")])
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys(username)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        usernameInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.EMPTY_ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == c.EMPTY_ERROR_MESSAGE_TEXT

    #case 2-Sadece şifre alanı boş geçildiğinde uyarı mesajı olarak "Epic sadface: Password is required" gösterilmelidir.
    @pytest.mark.parametrize("username, password", [("standard_user","")])
    def test_password_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys(username)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        usernameInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")))
        assert errorMessage.text == "Epic sadface: Password is required"


    #case 3-Kullanıcı adı "locked_out_user" şifre alanı "secret_sauce" gönderildiğinde "Epic sadface: Sorry, this user has been locked out." mesajı gösterilmelidir.
    def test_locked_user(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("locked_out_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")))
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."

    #case 4-Kullanıcı adı "standard_user" şifre "secret_sauce" gönderildiğinde kullanıcı "/inventory.html" sayfasına gönderilmelidir. Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır.
    def test_successful_login(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        usernameInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        current_Url = self.driver.current_url
        expected_Url = "https://www.saucedemo.com/inventory.html"
        if current_Url == expected_Url:
            print("Doğru sayfadasınız:", current_Url)
        else:
            print("Yanlış sayfaya yönlendirildiniz. Şu anki URL:"), current_Url

        listOfProducts = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, "inventory_item")))
        assert len(listOfProducts) == 6



 #Kullandığımız SauceDemo sitesinde kendi belirlediğiniz en az "3" test case daha yazınız.
#En az 1 testiniz parametrize fonksiyonu ile en az 3 farklı veriyle test edilmelidir.
# caseler;
# 1-Kullanıcı başarılı bir şekilde çıkış yapabilmeli.
# 2-error_user kullanıcı adıyla giriş yapıldığında ürünü sepetten kaldırma işlemi kontrolü
# 3-Verilen 3 farklı veri ile kullanıcı girişi
        

class Test_SauceClass:

    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get("https://www.saucedemo.com")
        self.driver.maximize_window() 

    def teardown_method(self): 
        self.driver.quit()


    def getData(sayfa):
        excel = openpyxl.load_workbook(c.LOGIN_XLSX)
        sheet = excel[sayfa] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data
    
    #Başarılı çıkış işlemi
    @pytest.mark.parametrize("username,password", getData("Sayfa1"))
    def test_Logout(self,username, password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        menuButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='react-burger-menu-btn']")))
        menuButton.click()
        logoutLink = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "logout_sidebar_link")))
        logoutLink.click()

    # error-user kullanıcı adı ile giriş yapıldığında ürünün sepetten kaldırılamama durumu
    @pytest.mark.parametrize("username,password", getData("Sayfa2"))
    def test_ErrorUser(self,username,password):
        self.driver.get("https://www.saucedemo.com/")
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='add-to-cart-sauce-labs-backpack']")))
        addToCart.click()
        removeButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='remove-sauce-labs-backpack']")))
        removeButton.click()
        errorRemove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='remove-sauce-labs-backpack']")))
        assert errorRemove.text == "Remove"


    #parametrize decorator ile 3 farklı veriyle giriş 
    @pytest.mark.parametrize("username,password", getData("Sayfa3"))
    def test_Login(self,username,password):
            usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
            usernameInput.send_keys(username)
            passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
            passwordInput.send_keys(password)
            loginButton = self.driver.find_element(By.ID, c.LOGIN_BUTTON_ID)
            loginButton.click()
            headerLogo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='header_container']/div[1]/div[2]/div")))
            assert headerLogo.text == "Swag Labs"

